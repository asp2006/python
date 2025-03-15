# هدف: نمایش نحوه کار با فایل‌های اکسل در پایتون با استفاده از openpyxl، شامل بارگذاری فایل، نوشتن داده‌ها، افزودن ردیف‌ها، فرمت‌دهی به سلول‌ها، ادغام سلول‌ها، افزودن فرمول‌ها و ایجاد نمودارها.


from openpyxl import load_workbook

# بارگذاری فایل اکسل
wb = load_workbook (r"C:\Users\kara\OneDrive\Documents\output_file.xlsx")
# دسترسی به شیت فعال (اولین شیت)w 
ws = wb.active
# نمایش داده‌های سلول‌ها
print(ws['A1'].value)  # نمایش مقدار سلول A1

# نوشتن داده به فایل اکسل :
from openpyxl import Workbook

# ساخت یک فایل اکسل جدید
wb = Workbook()
# دسترسی به شیت فعال
ws = wb.active

# نوشتن داده در سلول‌ها
ws['A1'] = 'Hello'
ws['B1'] = 'World'

# ذخیره فایل اکسل
wb.save(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")

# بارگذاری یک فایل اکسل موجود
wb = load_workbook(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")

# دسترسی به شیت خاص
ws = wb['Sheet']

# ویرایش داده در سلول A1
ws['A1'] = 'Hello'

# ذخیره تغییرات
wb.save(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")

# اضافه کردن ردیف‌ها و ستون‌ها :
# بارگذاری فایل اکسل
wb = load_workbook(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")
ws = wb.active

# افزودن یک ردیف جدید به شیت
ws.append([1, 2, 3, 4])

# ذخیره تغییرات
wb.save(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")

# فرمت‌دهی به سلول‌ها :

from openpyxl.styles import Font

# بارگذاری فایل اکسل
wb = load_workbook(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")
ws = wb.active

# تغییر فونت سلول A1
ws['A1'].font = Font(bold=True, color="FF0000")  # فونت بولد و رنگ قرمز

# ذخیره تغییرات
wb.save(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")

# تغییر رنگ پس‌زمینه سلول :

from openpyxl.styles import PatternFill

# تغییر رنگ پس‌زمینه سلول A1
ws['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# ذخیره تغییرات
wb.save(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")

# ادغام سلول‌ها :

# بارگذاری فایل اکسل
wb = load_workbook(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")
ws = wb.active

# ادغام سلول‌ها از A1 تا C1
ws.merge_cells('A1:C1')

# ذخیره تغییرات
wb.save(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")

#  افزودن فرمول‌ها به سلول‌ها :
# بارگذاری فایل اکسل
wb = load_workbook(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")
ws = wb.active

# افزودن فرمول به سلول A2
ws['A2'] = '=SUM(B1:B10)'

# ذخیره تغییرات
wb.save(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")

# ایجاد گراف‌ها و نمودارها :
from openpyxl.chart import BarChart, Reference
from openpyxl import load_workbook

# بارگذاری فایل اکسل
wb = load_workbook(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")

# دسترسی به شیت فعال
ws = wb.active

# ساخت یک نمودار میله‌ای
chart = BarChart()

# تنظیم داده‌ها برای نمودار
# می‌خواهیم داده‌ها از ستون دوم (B) و از ردیف 1 تا 10 را برای نمودار انتخاب کنیم
data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=10)

# افزودن داده‌ها به نمودار
# 'titles_from_data=True' به این معناست که عنوان نمودار از داده‌های موجود گرفته شود
chart.add_data(data, titles_from_data=True)

# اضافه کردن نمودار به شیت
# موقعیت نمودار را در سلول "E5" قرار می‌دهیم
ws.add_chart(chart, "E5")

# ذخیره تغییرات در فایل
wb.save(r"C:\Users\kara\OneDrive\Desktop\python\ASP\new_output.xlsx")
