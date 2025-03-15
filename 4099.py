# هدف: بارگذاری داده‌ها از چندین فایل اکسل و درج آن‌ها در کالکشن‌های مختلف در MongoDB با استفاده از نام فایل‌ها به عنوان نام کالکشن‌ها.

from openpyxl import load_workbook
from pymongo import MongoClient

# نام‌های فایل‌های اکسل
excel_files = [
    (r"C:\Users\kara\OneDrive\Desktop\python\ASP\ex1.xlsx"),
    (r"C:\Users\kara\OneDrive\Desktop\python\ASP\ex2.xlsx"),
    (r"C:\Users\kara\OneDrive\Desktop\python\ASP\ex3.xlsx"),
    (r"C:\Users\kara\OneDrive\Desktop\python\ASP\ex4.xlsx"),
]

# اتصال به MongoDB
client = MongoClient('mongodb://localhost:27017/')
db = client['my_new_database']

# پردازش هر فایل اکسل
for excel_file_path in excel_files:
    # بارگذاری فایل اکسل
    workbook = load_workbook(filename=excel_file_path)
    sheet = workbook.active

    # انتخاب نام کالکشن به نام فایل اکسل
    collection_name = excel_file_path.split('.')[0]  # استفاده از نام فایل به عنوان نام کالکشن
    collection = db[collection_name]

    # خواندن داده‌ها از اکسل و درج در MongoDB
    for row in sheet.iter_rows(min_row=2, max_row=11, min_col=1, max_col=5, values_only=True):
        data_dict = {
            'column1': row[0],
            'column2': row[1],
            'column3': row[2],
            'column4': row[3],
            'column5': row[4]
        }
        collection.insert_one(data_dict)

    print(f"داده‌ها از فایل {excel_file_path} با موفقیت در MongoDB ذخیره شدند.")

# بستن ارتباط با MongoDB
client.close()
